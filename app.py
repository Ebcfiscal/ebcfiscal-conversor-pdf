"""Conversor de estados de cuenta bancarios mexicanos de PDF a Excel.

Ejecucion local:
    streamlit run app.py

Dependencias:
    streamlit pdfplumber pandas openpyxl
"""

from __future__ import annotations

import io
import os
import re
import smtplib
import ssl
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from typing import Callable, Sequence

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BANKS = [
    "BBVA",
    "Banorte",
    "Santander",
    "Citibanamex",
    "Scotiabank",
    "HSBC",
    "Banca Afirme",
    "Banregio",
    "Banco Inbursa",
    "Banco Azteca",
    "Banco del Bajío",
    "BanCoppel",
    "Banco Bancrea",
    "Banco Mifel",
    "Banco Actinver",
]
COLUMNS = ["Fecha", "Concepto / Descripción", "Depósito", "Retiro", "Saldo"]
EMAIL_RE = re.compile(r"^[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Z0-9-]+(?:\.[A-Z0-9-]+)+$", re.I)

DATE_TOKEN = r"(?:\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?|\d{1,2}(?:\s+|[/-])(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|SEPT|OCT|NOV|DIC)[A-Z]*\.?(?:(?:\s+|[/-])\d{2,4})?)"
MONEY_TOKEN = r"(?:\(\s*)?(?:\$\s*)?-?\s*\d[\d,]*(?:\.\d{2})(?:\s*\))?-?"
DATE_RE = re.compile(rf"^\s*(?P<date>{DATE_TOKEN})\b", re.I)
DATE_ANY_RE = re.compile(rf"(?P<date>{DATE_TOKEN})\b", re.I)
MONEY_RE = re.compile(rf"(?<!\w)({MONEY_TOKEN})(?!\w)", re.I)

MONTHS = {
    "ENE": 1, "ENERO": 1, "FEB": 2, "FEBRERO": 2, "MAR": 3, "MARZO": 3,
    "ABR": 4, "ABRIL": 4, "MAY": 5, "MAYO": 5, "JUN": 6, "JUNIO": 6,
    "JUL": 7, "JULIO": 7, "AGO": 8, "AGOSTO": 8, "SEP": 9,
    "SEPT": 9, "SEPTIEMBRE": 9, "OCT": 10, "OCTUBRE": 10,
    "NOV": 11, "NOVIEMBRE": 11, "DIC": 12, "DICIEMBRE": 12,
}


@dataclass(frozen=True)
class BankConfig:
    """Pistas de formato para interpretar los movimientos de cada banco."""

    headers: tuple[str, ...]
    ignore: tuple[str, ...]
    amount_order: tuple[str, ...]
    deposit_words: tuple[str, ...]
    withdrawal_words: tuple[str, ...]


COMMON_IGNORE = (
    "saldo anterior", "saldo inicial", "total depositos", "total depósitos",
    "total retiros", "total cargos", "total abonos", "retiros totales",
    "cargos totales", "depositos totales", "depósitos totales", "abonos totales",
    "resumen", "pagina ", "página ", "fecha concepto",
    "fecha descripcion", "fecha descripción", "estado de cuenta", "periodo",
)

SUMMARY_TOTAL_RE = re.compile(
    r"\b(?:total(?:es)?|suma(?:s)?|acumulado(?:s)?)"
    r"(?:\s+(?:de|del|los|las|en|el|periodo|mes))*\s+"
    r"(?:retiros?|cargos?|depositos?|abonos?|movimientos?)\b|"
    r"\b(?:retiros?|cargos?|depositos?|abonos?|movimientos?)\s+"
    r"(?:totales?|acumulados?)\b",
    re.I,
)

CONFIGS = {
    "BBVA": BankConfig(
        ("FECHA", "OPERACIÓN", "LIQUIDACIÓN", "CONCEPTO", "CARGO", "ABONO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "spei recibido", "traspaso recibido"),
        ("cargo", "retiro", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "Banorte": BankConfig(
        ("FECHA", "DESCRIPCIÓN", "DEPÓSITOS", "RETIROS", "SALDO"),
        COMMON_IGNORE, ("Depósito", "Retiro", "Saldo"),
        ("deposito", "depósito", "abono", "transferencia recibida", "spei recibido"),
        ("retiro", "cargo", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "Santander": BankConfig(
        ("FECHA", "CONCEPTO", "RETIRO", "DEPÓSITO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("deposito", "depósito", "abono", "transferencia recibida"),
        ("retiro", "cargo", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "Citibanamex": BankConfig(
        ("FECHA", "CONCEPTO", "RETIROS", "DEPÓSITOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("deposito", "depósito", "abono", "transferencia recibida"),
        ("retiro", "cargo", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "Scotiabank": BankConfig(
        ("FECHA", "CONCEPTO", "CARGOS", "ABONOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "transferencia recibida"),
        ("cargo", "retiro", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "HSBC": BankConfig(
        ("FECHA", "DESCRIPCIÓN", "DEPÓSITO/ABONO", "RETIRO/CARGO", "SALDO"),
        COMMON_IGNORE, ("Depósito", "Retiro", "Saldo"),
        ("abono", "deposito", "depósito", "transferencia recibida"),
        ("cargo", "retiro", "compra", "comision", "comisión", "pago", "spei enviado", "transferencia enviada"),
    ),
    "Banca Afirme": BankConfig(
        ("FECHA", "REFERENCIA", "CONCEPTO", "CARGOS", "ABONOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "traspaso recibido"),
        ("cargo", "retiro", "compra", "comision", "comisión", "pago", "spei enviado", "traspaso enviado"),
    ),
    "Banregio": BankConfig(
        ("FECHA", "DESCRIPCIÓN", "DEPÓSITOS", "RETIROS", "SALDO"),
        COMMON_IGNORE, ("Depósito", "Retiro", "Saldo"),
        ("deposito", "depósito", "abono", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("retiro", "cargo", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco Inbursa": BankConfig(
        ("FECHA", "CONCEPTO", "DESCRIPCIÓN", "CARGO", "ABONO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco Azteca": BankConfig(
        ("FECHA", "MOVIMIENTO", "DESCRIPCIÓN", "CARGO", "ABONO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "pago recibido"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco del Bajío": BankConfig(
        ("FECHA", "REFERENCIA", "DESCRIPCIÓN", "CARGOS", "ABONOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "BanCoppel": BankConfig(
        ("FECHA", "DESCRIPCIÓN", "CONCEPTO", "CARGO", "ABONO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco Bancrea": BankConfig(
        ("FECHA", "OPERACIÓN", "CONCEPTO", "CARGO", "ABONO", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco Mifel": BankConfig(
        ("FECHA", "REFERENCIA", "DESCRIPCIÓN", "CARGOS", "ABONOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "transferencia recibida"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado"),
    ),
    "Banco Actinver": BankConfig(
        ("FECHA", "MOVIMIENTO", "CONCEPTO", "RETIROS", "DEPÓSITOS", "SALDO"),
        COMMON_IGNORE, ("Retiro", "Depósito", "Saldo"),
        ("abono", "deposito", "depósito", "credito", "crédito", "spei recibido", "venta", "liquidación a favor"),
        ("cargo", "retiro", "debito", "débito", "compra", "comision", "comisión", "pago", "spei enviado", "compra de títulos"),
    ),
}


def clean_text(value: object) -> str:
    """Normaliza espacios y saltos de linea sin eliminar acentos."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def comparable(value: str) -> str:
    value = unicodedata.normalize("NFKD", clean_text(value)).encode("ascii", "ignore").decode()
    return value.casefold()


def is_summary_total(value: str) -> bool:
    """Detecta renglones de totales aunque la etiqueta quede en otra columna."""
    normalized = comparable(value)
    if SUMMARY_TOTAL_RE.search(normalized):
        return True
    # Santander cierra su tabla con una etiqueta genérica, por ejemplo:
    # "TOTAL 1,610,000.00 1,360,783.00", sin escribir depósitos/retiros.
    return bool(
        re.match(r"^total(?:es)?\b", normalized)
        and MONEY_RE.search(clean_text(value))
    )


def money_to_float(value: object) -> float:
    """Convierte $ 1,234.56, (1,234.56) o -1,234.56 a float."""
    if value is None or clean_text(value) in {"", "-", "—"}:
        return 0.0
    raw = clean_text(value)
    negative = (raw.startswith("(") and raw.endswith(")")) or raw.rstrip().endswith("-")
    raw = re.sub(r"[^\d.\-]", "", raw.replace(",", ""))
    if raw.endswith("-"):
        raw = raw[:-1]
    try:
        amount = float(raw)
    except ValueError:
        return 0.0
    return round(-abs(amount) if negative else amount, 2)


def infer_statement_year(lines: Sequence[str]) -> int:
    """Busca el anio del periodo; usa el actual solo como ultimo recurso."""
    period_re = re.compile(r"(?:PERIODO|CORTE|AL)\D{0,25}(20\d{2})", re.I)
    for line in lines:
        match = period_re.search(line)
        if match:
            return int(match.group(1))
    years = [int(y) for line in lines[:80] for y in re.findall(r"\b(20\d{2})\b", line)]
    return max(set(years), key=years.count) if years else datetime.now().year


def infer_statement_month(lines: Sequence[str]) -> int | None:
    """Obtiene el mes del período/corte para bancos que imprimen sólo el día."""
    prioritized = [line for line in lines if any(token in comparable(line) for token in ("periodo", "corte"))]
    candidates = prioritized + list(lines[:100])
    month_names = "|".join(sorted(MONTHS, key=len, reverse=True))
    for line in candidates:
        normalized = comparable(line).upper()
        named = re.search(rf"\b({month_names})\b", normalized)
        if named:
            return MONTHS[named.group(1)]
        numeric = re.search(r"\b\d{1,2}[/-](\d{1,2})[/-]20\d{2}\b", normalized)
        if numeric and 1 <= int(numeric.group(1)) <= 12:
            return int(numeric.group(1))
    return None


def normalize_date(value: str, default_year: int, default_month: int | None = None) -> str:
    """Devuelve una fecha bancaria en formato DD/MM/YYYY."""
    text = comparable(value).upper().replace(".", "")
    if re.fullmatch(r"\d{1,2}", text):
        if default_month is None:
            raise ValueError(f"Falta el mes para la fecha: {value}")
        return datetime(default_year, default_month, int(text)).strftime("%d/%m/%Y")
    numeric = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", text)
    if numeric:
        day, month = int(numeric.group(1)), int(numeric.group(2))
        year = int(numeric.group(3)) if numeric.group(3) else default_year
        year += 2000 if year < 100 else 0
        return datetime(year, month, day).strftime("%d/%m/%Y")
    text = re.sub(r"[/-]", " ", text)
    text = clean_text(text)
    named = re.fullmatch(r"(\d{1,2})\s+([A-Z]+)(?:\s+(\d{2,4}))?", text)
    if not named or named.group(2) not in MONTHS:
        raise ValueError(f"Fecha no reconocida: {value}")
    year = int(named.group(3)) if named.group(3) else default_year
    year += 2000 if year < 100 else 0
    return datetime(year, MONTHS[named.group(2)], int(named.group(1))).strftime("%d/%m/%Y")


def normalize_day_month(value: str) -> str:
    """Normaliza una fecha sin año a DD/MM, como la fecha LIQ de BBVA."""
    text = comparable(value).upper().replace(".", "")
    numeric = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})(?:[/-]\d{2,4})?", text)
    if numeric:
        day, month = int(numeric.group(1)), int(numeric.group(2))
    else:
        named_text = clean_text(re.sub(r"[/-]", " ", text))
        named = re.fullmatch(r"(\d{1,2})\s+([A-Z]+)(?:\s+\d{2,4})?", named_text)
        if not named or named.group(2) not in MONTHS:
            raise ValueError(f"Fecha sin año no reconocida: {value}")
        day, month = int(named.group(1)), MONTHS[named.group(2)]
    # El año 2000 permite validar también 29/FEB sin incorporarlo al resultado.
    datetime(2000, month, day)
    return f"{day:02d}/{month:02d}"


def group_words_by_line(words: Sequence[dict], tolerance: float = 3.0) -> list[list[dict]]:
    """Agrupa palabras por su coordenada vertical."""
    if not words:
        return []
    rows: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (round(float(w["top"]) / tolerance), float(w["x0"]))):
        if not rows or abs(float(word["top"]) - sum(float(w["top"]) for w in rows[-1]) / len(rows[-1])) > tolerance:
            rows.append([word])
        else:
            rows[-1].append(word)
    return [sorted(row, key=lambda item: float(item["x0"])) for row in rows]


def words_to_lines(words: Sequence[dict], tolerance: float = 3.0) -> list[str]:
    """Reconstruye líneas por coordenadas para el respaldo de texto corrido."""
    return [clean_text(" ".join(w["text"] for w in row)) for row in group_words_by_line(words, tolerance)]


def word_center(word: dict) -> float:
    return (float(word["x0"]) + float(word.get("x1", word["x0"]))) / 2


def coordinate_header_anchors(rows: Sequence[Sequence[dict]], start: int) -> tuple[dict[str, float], int] | None:
    """Localiza encabezados aunque estén repartidos en dos líneas visuales."""
    for span in (1, 2):
        if start + span > len(rows):
            continue
        candidate = [word for row in rows[start:start + span] for word in row]
        positions: dict[str, list[float]] = {}
        for word in candidate:
            role = header_role(str(word.get("text", "")))
            if role:
                positions.setdefault(role, []).append(word_center(word))
        if "date" not in positions:
            continue
        has_direction = "deposit" in positions or "withdrawal" in positions
        if not has_direction and "amount" not in positions:
            continue
        if not ("balance" in positions or {"deposit", "withdrawal"}.issubset(positions)):
            continue
        anchors: dict[str, float] = {}
        for role, values in positions.items():
            if role == "date":
                anchors[role] = min(values)
            elif role == "amount" and len(values) > 1:
                # Un encabezado genérico repetido no permite distinguir columnas.
                continue
            else:
                anchors[role] = sum(values) / len(values)
        if "description" not in anchors:
            monetary_x = [anchors[role] for role in ("deposit", "withdrawal", "amount", "balance") if role in anchors]
            if monetary_x:
                anchors["description"] = (anchors["date"] + min(monetary_x)) / 2
        return anchors, start + span
    return None


def coordinate_table_rows(words: Sequence[dict]) -> list[list[str]]:
    """Crea una tabla a partir de posiciones x/y cuando el PDF no contiene celdas."""
    visual_rows = group_words_by_line(words)
    result: list[list[str]] = []
    index = 0
    anchors: dict[str, float] | None = None
    data_start = 0

    while index < len(visual_rows):
        located = coordinate_header_anchors(visual_rows, index)
        if located:
            anchors, data_start = located
            result.append(["FECHA", "DESCRIPCION", "DEPOSITOS", "RETIROS", "SALDO", "IMPORTE"])
            index = data_start
            break
        index += 1
    if not anchors:
        return []

    monetary_roles = [role for role in ("deposit", "withdrawal", "balance", "amount") if role in anchors]
    first_monetary_x = min(anchors[role] for role in monetary_roles)
    description_limit = (anchors["description"] + first_monetary_x) / 2
    monetary_positions = sorted(anchors[role] for role in monetary_roles)
    gaps = [right - left for left, right in zip(monetary_positions, monetary_positions[1:]) if right > left]
    max_anchor_distance = max(38.0, min(80.0, (min(gaps) * 0.68) if gaps else 65.0))

    for visual_row in visual_rows[data_start:]:
        repeated = coordinate_header_anchors([visual_row], 0)
        if repeated:
            new_anchors, _ = repeated
            anchors.update(new_anchors)
            monetary_roles = [role for role in ("deposit", "withdrawal", "balance", "amount") if role in anchors]
            first_monetary_x = min(anchors[role] for role in monetary_roles)
            description_limit = (anchors["description"] + first_monetary_x) / 2
            monetary_positions = sorted(anchors[role] for role in monetary_roles)
            gaps = [right - left for left, right in zip(monetary_positions, monetary_positions[1:]) if right > left]
            max_anchor_distance = max(38.0, min(80.0, (min(gaps) * 0.68) if gaps else 65.0))
            continue

        line = clean_text(" ".join(str(word.get("text", "")) for word in visual_row))
        # El total mensual marca el fin del detalle. Detener la lectura evita
        # que sus importes y las leyendas posteriores se adhieran al último
        # movimiento cuando el renglón no contiene fecha.
        if is_summary_total(line):
            break
        date_match = DATE_ANY_RE.search(line)
        if date_match and date_match.start() > 24:
            date_match = None
        day_only_index: int | None = None
        if not date_match:
            for candidate_index, word in enumerate(visual_row):
                candidate_text = clean_text(word.get("text", ""))
                if re.fullmatch(r"\d{1,2}", candidate_text) and abs(word_center(word) - anchors["date"]) <= 55:
                    day_only_index = candidate_index
                    break

        role_values: dict[str, list[str]] = {role: [] for role in monetary_roles}
        monetary_word_ids: set[int] = set()
        for word_index, word in enumerate(visual_row):
            text = clean_text(word.get("text", ""))
            if not MONEY_RE.fullmatch(text):
                continue
            x = word_center(word)
            if x < description_limit:
                continue
            nearest_role = min(monetary_roles, key=lambda role: abs(x - anchors[role]))
            anchor_distance = abs(x - anchors[nearest_role])
            if anchor_distance > max_anchor_distance:
                continue
            context = " ".join(
                comparable(clean_text(visual_row[neighbor].get("text", "")))
                for neighbor in range(max(0, word_index - 2), min(len(visual_row), word_index + 3))
                if neighbor != word_index
            )
            foreign_context = any(
                marker in context
                for marker in ("usd", "dls", "dlls", "dolar", "dolares", "eur", "euro", "cambio", "%")
            )
            # Una cifra perfectamente alineada con una columna monetaria prevalece;
            # una cifra desplazada y rodeada de USD/tipo de cambio pertenece al concepto.
            if foreign_context and anchor_distance > 28:
                continue
            role_values[nearest_role].append(text)
            monetary_word_ids.add(word_index)

        if date_match or day_only_index is not None:
            date_text = date_match.group("date") if date_match else clean_text(visual_row[day_only_index].get("text", ""))
            date_word_count = day_only_index + 1 if day_only_index is not None else 0
            attached_description = ""
            if date_match:
                # Algunos PDF de Banorte unen fecha y concepto en una sola
                # palabra interna: "26-ENE-26DEP.EFECTIVO". Conservamos el
                # sufijo para que el movimiento no quede sin descripción.
                cursor = 0
                for word_index, word in enumerate(visual_row):
                    word_text = clean_text(word.get("text", ""))
                    word_start = cursor
                    word_end = word_start + len(word_text)
                    if word_start <= date_match.end() <= word_end:
                        date_word_count = word_index + 1
                        attached_description = clean_text(word_text[date_match.end() - word_start:])
                        break
                    cursor = word_end + 1
                year_suffix = re.match(r"^([/-]\d{2,4})(?=\D|$)", attached_description)
                if year_suffix and not re.search(r"[/-]\d{2,4}$", date_text):
                    date_text = f"{date_text}{year_suffix.group(1)}"
                    attached_description = clean_text(attached_description[year_suffix.end():])
            description_words = [
                clean_text(word.get("text", ""))
                for word_index, word in enumerate(visual_row)
                if word_index >= date_word_count and word_index not in monetary_word_ids
            ]
            if attached_description:
                description_words.insert(0, attached_description)
            description = clean_text(" ".join(description_words))
            if any(role_values.values()):
                result.append([
                    date_text,
                    description,
                    clean_text(" ".join(role_values.get("deposit", []))),
                    clean_text(" ".join(role_values.get("withdrawal", []))),
                    clean_text(" ".join(role_values.get("balance", []))),
                    clean_text(" ".join(role_values.get("amount", []))),
                ])
        elif result and any(role_values.values()):
            # Afirme y otros bancos pueden imprimir el saldo varias líneas
            # después del depósito/retiro del mismo movimiento.
            target = next((item for item in reversed(result[1:]) if item[0]), None)
            if target:
                role_columns = {"deposit": 2, "withdrawal": 3, "balance": 4, "amount": 5}
                for role, values in role_values.items():
                    if values and not target[role_columns[role]]:
                        target[role_columns[role]] = clean_text(" ".join(values))
        elif result:
            continuation = clean_text(" ".join(
                clean_text(word.get("text", ""))
                for word in visual_row
                if word_center(word) < first_monetary_x
            ))
            if continuation:
                result.append(["", continuation, "", "", "", ""])
    return result


def extract_pdf_rows(pdf_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    """Extrae texto posicional y tablas delineadas de todas las paginas."""
    lines: list[str] = []
    table_rows: list[list[str]] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines.extend(words_to_lines(page_words))
            # Muchos bancos dibujan columnas visuales sin bordes de tabla.
            # Esta ruta conserva cargos/abonos usando las coordenadas x/y.
            coordinate_rows = coordinate_table_rows(page_words)
            if coordinate_rows:
                table_rows.extend(coordinate_rows)
                continue
            settings_candidates = (
                {"vertical_strategy": "lines", "horizontal_strategy": "lines", "snap_tolerance": 3},
                {"vertical_strategy": "text", "horizontal_strategy": "text", "intersection_tolerance": 5},
            )
            page_tables: list[list[list[str | None]]] = []
            for settings in settings_candidates:
                try:
                    detected = page.extract_tables(table_settings=settings) or []
                except Exception:  # Algunos PDFs tienen trazos malformados.
                    detected = []
                page_tables.extend(detected)
            for table in page_tables:
                for row in table:
                    cleaned = [clean_text(cell) for cell in (row or [])]
                    if any(cleaned):
                        table_rows.append(cleaned)
    return [line for line in lines if line], table_rows


def is_noise(line: str, config: BankConfig) -> bool:
    normalized = comparable(line)
    if is_summary_total(normalized):
        return True
    if any(token in normalized for token in map(comparable, config.ignore)):
        return True
    header_hits = sum(comparable(header) in normalized for header in config.headers)
    return header_hits >= 3


def table_rows_to_lines(rows: Sequence[Sequence[str]]) -> list[str]:
    return [clean_text(" ".join(cell for cell in row if clean_text(cell))) for row in rows]


def header_role(value: str) -> str | None:
    """Clasifica el encabezado de una columna bancaria sin depender de acentos."""
    text = comparable(value)
    if not text:
        return None
    deposit_terms = ("deposito", "depositos", "abono", "abonos", "credito")
    withdrawal_terms = ("retiro", "retiros", "cargo", "cargos", "debito")
    has_deposit_term = any(word in text for word in deposit_terms)
    has_withdrawal_term = any(word in text for word in withdrawal_terms)
    if "fecha" in text or text in {"dia", "f operacion", "f. operacion"}:
        return "date"
    if has_deposit_term and has_withdrawal_term:
        return "amount"
    if has_deposit_term:
        return "deposit"
    if has_withdrawal_term:
        return "withdrawal"
    if "saldo" in text:
        return "balance"
    if any(word in text for word in ("importe", "monto")):
        return "amount"
    if any(word in text for word in ("concepto", "descripcion", "detalle", "movimiento", "operacion", "referencia")):
        return "description"
    return None


def locate_table_header(row: Sequence[str]) -> dict[str, list[int]] | None:
    """Devuelve las posiciones de columnas sólo si la fila parece un encabezado real."""
    mapping: dict[str, list[int]] = {}
    for index, cell in enumerate(row):
        role = header_role(clean_text(cell))
        if role:
            mapping.setdefault(role, []).append(index)
    monetary_roles = {"deposit", "withdrawal", "balance", "amount"}.intersection(mapping)
    if "date" not in mapping or not monetary_roles:
        return None
    if not ({"deposit", "withdrawal"}.intersection(mapping) or "amount" in mapping):
        return None
    return mapping


def cell_amount(value: object) -> float | None:
    """Lee un importe sólo cuando la celda contiene un valor monetario explícito."""
    matches = list(MONEY_RE.finditer(clean_text(value)))
    if not matches:
        return None
    return money_to_float(matches[-1].group(1))


def make_frame(records: Sequence[dict]) -> pd.DataFrame:
    """Valida los movimientos y elimina filas sin una operación monetaria comprobable."""
    frame = pd.DataFrame(records, columns=COLUMNS)
    if frame.empty:
        return frame
    frame["Concepto / Descripción"] = frame["Concepto / Descripción"].map(clean_text)
    for column in ("Depósito", "Retiro"):
        frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0.0).abs().round(2)
    frame["Saldo"] = pd.to_numeric(frame["Saldo"], errors="coerce").round(2)
    has_movement = frame["Depósito"].gt(0) | frame["Retiro"].gt(0)
    frame = frame[has_movement & frame["Concepto / Descripción"].ne("")]
    return frame.drop_duplicates().reset_index(drop=True)


def extract_declared_totals(lines: Sequence[str]) -> dict[str, float]:
    """Lee totales del resumen sin confundir encabezados o movimientos."""
    candidates = declared_total_candidates(lines)
    totals: dict[str, float] = {}
    for column, values in candidates.items():
        if values:
            # El total mensual es normalmente el mayor de los subtotales; las
            # leyendas repetidas en el PDF cuentan una sola vez.
            totals[column] = round(max(values), 2)
    return totals


def declared_total_candidates(lines: Sequence[str]) -> dict[str, list[float]]:
    """Obtiene importes únicos impresos como totales de depósitos o retiros."""
    candidates: dict[str, list[float]] = {"Depósito": [], "Retiro": []}
    for line in lines:
        normalized = comparable(line)
        if "deposito" in normalized and "retiro" in normalized:
            continue
        amounts = [abs(money_to_float(match.group(1))) for match in MONEY_RE.finditer(line)]
        if not amounts:
            continue
        if re.search(r"\b(?:depositos?|abonos?)\b", normalized) and is_summary_total(normalized):
            candidates["Depósito"].append(round(amounts[-1], 2))
        elif re.search(r"\b(?:retiros?|cargos?)\b", normalized) and is_summary_total(normalized):
            candidates["Retiro"].append(round(amounts[-1], 2))
    return {
        column: list(dict.fromkeys(values))
        for column, values in candidates.items()
    }


def santander_declared_total_candidates(lines: Sequence[str]) -> dict[str, list[float]]:
    """Lee el TOTAL genérico de Santander: primero depósito y después retiro."""
    candidates = declared_total_candidates(lines)
    for line in lines:
        if not re.match(r"^total(?:es)?\b", comparable(line)):
            continue
        amounts = [abs(money_to_float(match.group(1))) for match in MONEY_RE.finditer(line)]
        if len(amounts) < 2:
            continue
        candidates["Depósito"].append(round(amounts[-2], 2))
        candidates["Retiro"].append(round(amounts[-1], 2))
    return {
        column: list(dict.fromkeys(values))
        for column, values in candidates.items()
    }


def remove_duplicated_declared_totals(frame: pd.DataFrame, lines: Sequence[str]) -> pd.DataFrame:
    """Elimina cifras de resumen que duplican exactamente el detalle Santander."""
    if frame.empty:
        return frame
    result = frame.copy()
    candidates = santander_declared_total_candidates(lines)
    tolerance = 0.03

    for column in ("Retiro", "Depósito"):
        values = pd.to_numeric(result[column], errors="coerce").fillna(0.0)
        column_total = round(float(values.sum()), 2)
        for declared in candidates[column]:
            # Una fila es un resumen duplicado sólo cuando contiene ella sola
            # el total declarado y todas las demás filas ya suman ese total.
            duplicate_indexes = [
                index for index, amount in values.items()
                if abs(float(amount) - declared) <= tolerance
                and abs((column_total - float(amount)) - declared) <= tolerance
            ]
            if not duplicate_indexes:
                continue
            duplicate_index = duplicate_indexes[-1]
            result.at[duplicate_index, column] = 0.0
            values.at[duplicate_index] = 0.0
            column_total = round(float(values.sum()), 2)
            break

    has_movement = result["Depósito"].gt(0) | result["Retiro"].gt(0)
    return result[has_movement].reset_index(drop=True)


def validate_extraction_totals(frame: pd.DataFrame, lines: Sequence[str]) -> list[str]:
    """Advierte si el detalle no concilia con los totales impresos por el banco."""
    declared = extract_declared_totals(lines)
    warnings: list[str] = []
    labels = {"Depósito": "depósitos", "Retiro": "retiros"}
    for column, expected in declared.items():
        extracted = round(float(frame[column].sum()), 2)
        tolerance = max(0.02, abs(expected) * 0.000001)
        if abs(extracted - expected) > tolerance:
            warnings.append(
                f"Los {labels[column]} extraídos suman ${extracted:,.2f}, "
                f"pero el resumen del PDF indica ${expected:,.2f}."
            )

    # Comprueba que cada saldo se explique por el movimiento correspondiente.
    balances = pd.to_numeric(frame["Saldo"], errors="coerce")
    dates = pd.to_datetime(frame["Fecha"], format="%d/%m/%Y", errors="coerce")
    failures: list[int] = []
    descending = len(dates) > 1 and dates.iloc[-1] < dates.iloc[0]
    for index in range(1, len(frame)):
        if pd.isna(balances.iloc[index - 1]) or pd.isna(balances.iloc[index]):
            continue
        if descending:
            expected_balance = (
                balances.iloc[index - 1]
                - float(frame["Depósito"].iloc[index - 1])
                + float(frame["Retiro"].iloc[index - 1])
            )
        else:
            expected_balance = (
                balances.iloc[index - 1]
                + float(frame["Depósito"].iloc[index])
                - float(frame["Retiro"].iloc[index])
            )
        if abs(float(balances.iloc[index]) - expected_balance) > 0.03:
            failures.append(index)
    if failures:
        first = failures[0]
        warnings.append(
            f"El saldo no concilia en {len(failures)} movimiento(s); "
            f"la primera diferencia aparece el {frame['Fecha'].iloc[first]}."
        )
    return warnings


def parse_structured_tables(
    rows: Sequence[Sequence[str]],
    config: BankConfig,
    default_year: int,
    default_month: int | None,
) -> pd.DataFrame:
    """Interpreta tablas conservando las celdas vacías de cargos y abonos."""
    records: list[dict] = []
    columns: dict[str, list[int]] | None = None
    current: dict | None = None
    previous_balance: float | None = None

    for raw_row in rows:
        row = [clean_text(cell) for cell in raw_row]
        detected_header = locate_table_header(row)
        if detected_header:
            if current:
                records.append(current)
                previous_balance = current.get("Saldo")
                current = None
            columns = detected_header
            continue
        if not columns:
            continue

        # Las filas creadas por coordenadas conservan el ancho completo de la
        # tabla. Una fila mucho más corta indica que terminó el detalle y que
        # comenzó un pie de página, texto legal u otra sección del estado.
        table_indexes = [index for indexes in columns.values() for index in indexes]
        required_width = (max(table_indexes) + 1) if table_indexes else 0
        if required_width and len(row) < required_width:
            if current:
                records.append(current)
                previous_balance = current.get("Saldo")
                current = None
            columns = None
            continue

        def joined(role: str) -> str:
            return clean_text(" ".join(row[index] for index in columns.get(role, []) if index < len(row)))

        date_cell = joined("date")
        date_match = DATE_RE.match(date_cell)
        day_only = re.fullmatch(r"\d{1,2}", date_cell)
        monetary_indexes = {
            index
            for role in ("deposit", "withdrawal", "balance", "amount")
            for index in columns.get(role, [])
        }
        date_indexes = set(columns.get("date", []))
        description_indexes = set(columns.get("description", []))
        if not description_indexes:
            description_indexes = set(range(len(row))) - monetary_indexes - date_indexes
        description = clean_text(" ".join(row[index] for index in sorted(description_indexes) if index < len(row)))

        if date_match or day_only:
            if current:
                records.append(current)
                previous_balance = current.get("Saldo")
                current = None
            # Los bancos suelen cerrar la tabla con renglones como
            # "TOTAL RETIROS" o "TOTAL DEPÓSITOS". Aunque alguna cifra quede
            # alineada con la columna de fecha, nunca representan movimientos.
            if is_noise(description, config):
                continue
            try:
                date_text = date_match.group("date") if date_match else date_cell
                date = normalize_date(date_text, default_year, default_month)
            except (ValueError, OverflowError):
                current = None
                continue

            deposit_raw = cell_amount(joined("deposit"))
            withdrawal_raw = cell_amount(joined("withdrawal"))
            balance = cell_amount(joined("balance"))
            generic_amount = cell_amount(joined("amount"))
            deposit = abs(deposit_raw) if deposit_raw is not None else 0.0
            withdrawal = abs(withdrawal_raw) if withdrawal_raw is not None else 0.0

            if generic_amount is not None and not deposit and not withdrawal:
                normalized = comparable(description)
                is_deposit = any(comparable(word) in normalized for word in config.deposit_words)
                is_withdrawal = any(comparable(word) in normalized for word in config.withdrawal_words)
                if generic_amount < 0 or (is_withdrawal and not is_deposit):
                    withdrawal = abs(generic_amount)
                elif is_deposit and not is_withdrawal:
                    deposit = abs(generic_amount)
                elif balance is not None and previous_balance is not None:
                    difference = round(balance - previous_balance, 2)
                    if abs(abs(difference) - abs(generic_amount)) <= 0.02:
                        if difference > 0:
                            deposit = abs(generic_amount)
                        elif difference < 0:
                            withdrawal = abs(generic_amount)

            current = {
                "Fecha": date,
                "Concepto / Descripción": description,
                "Depósito": round(deposit, 2),
                "Retiro": round(withdrawal, 2),
                "Saldo": balance,
            }
        elif current and description and not any(cell_amount(row[index]) is not None for index in monetary_indexes if index < len(row)):
            # Una continuación es válida sólo dentro de la misma tabla y sin importes.
            if not is_noise(description, config):
                current["Concepto / Descripción"] = clean_text(
                    f"{current['Concepto / Descripción']} {description}"
                )

    if current:
        records.append(current)
    return make_frame(records)


def classify_amounts(description: str, amounts: list[float], config: BankConfig) -> tuple[float, float, float | None]:
    """Mapea importes a deposito/retiro/saldo usando columnas y pistas semanticas."""
    if not amounts:
        return 0.0, 0.0, None
    normalized = comparable(description)
    is_deposit = any(comparable(word) in normalized for word in config.deposit_words)
    is_withdrawal = any(comparable(word) in normalized for word in config.withdrawal_words)
    deposit, withdrawal, balance = 0.0, 0.0, None

    if len(amounts) >= 3:
        values: dict[str, float] = {}
        for label, amount in zip(reversed(config.amount_order), reversed(amounts)):
            values[label] = abs(amount)
        deposit = values.get("Depósito", 0.0)
        withdrawal = values.get("Retiro", 0.0)
        balance = values.get("Saldo")
    elif len(amounts) == 2:
        # En texto corrido el primer valor es la operación y el último el saldo.
        # La dirección sólo se acepta si el concepto aporta evidencia; no se adivina.
        transaction, balance = abs(amounts[0]), amounts[1]
        if amounts[0] < 0 or (is_withdrawal and not is_deposit):
            withdrawal = transaction
        elif is_deposit and not is_withdrawal:
            deposit = transaction
    elif len(amounts) == 1:
        transaction = abs(amounts[0])
        if amounts[0] < 0 or (is_withdrawal and not is_deposit):
            withdrawal = transaction
        elif is_deposit and not is_withdrawal:
            deposit = transaction
    return round(deposit, 2), round(withdrawal, 2), None if balance is None else round(balance, 2)


def parse_bank(lines: Sequence[str], table_rows: Sequence[Sequence[str]], config: BankConfig) -> pd.DataFrame:
    """Usa primero tablas con columnas; recurre a texto sólo de forma estricta."""
    table_lines = table_rows_to_lines(table_rows)
    year = infer_statement_year(list(lines) + table_lines)
    month = infer_statement_month(list(lines) + table_lines)
    structured = parse_structured_tables(table_rows, config, year, month)
    if not structured.empty:
        return structured

    candidates = list(lines)
    if sum(bool(DATE_RE.match(line)) for line in table_lines) > sum(bool(DATE_RE.match(line)) for line in candidates):
        candidates = table_lines

    records: list[dict] = []
    current: dict | None = None
    continuation_markers = (
        "referencia", "ref ", "folio", "clave", "rastreo", "beneficiario",
        "ordenante", "rfc", "spei", "concepto", "cuenta destino",
    )

    for raw_line in candidates:
        line = clean_text(raw_line)
        match = DATE_RE.match(line)
        if match and not is_noise(line, config):
            if current:
                records.append(current)
            date_text = match.group("date")
            remainder = clean_text(line[match.end():])
            money_matches = list(MONEY_RE.finditer(remainder))
            amounts = [money_to_float(item.group(1)) for item in money_matches]
            description = remainder[: money_matches[0].start()] if money_matches else remainder
            deposit, withdrawal, balance = classify_amounts(description, amounts, config)
            try:
                normalized_date = normalize_date(date_text, year)
            except (ValueError, OverflowError):
                current = None
                continue
            current = {
                "Fecha": normalized_date,
                "Concepto / Descripción": clean_text(description),
                "Depósito": deposit,
                "Retiro": withdrawal,
                "Saldo": balance,
            }
        elif current and line and not is_noise(line, config) and not MONEY_RE.search(line):
            continuation = clean_text(line)
            normalized_continuation = comparable(continuation)
            if continuation and any(marker in normalized_continuation for marker in continuation_markers):
                current["Concepto / Descripción"] = clean_text(
                    f"{current['Concepto / Descripción']} {continuation}"
                )
    if current:
        records.append(current)

    return make_frame(records)


def prepare_bbva_liquidation_rows(rows: Sequence[Sequence[str]]) -> list[list[str]]:
    """Marca LIQ como la única columna de fecha en tablas BBVA con OPER y LIQ."""
    prepared: list[list[str]] = []
    for raw_row in rows:
        row = [clean_text(cell) for cell in raw_row]
        normalized = [comparable(cell) for cell in row]
        liquidation_indexes = [
            index
            for index, cell in enumerate(normalized)
            if cell in {"liq", "liquidacion", "fecha liq", "fecha liquidacion"}
            or "fecha liq" in cell
        ]
        if liquidation_indexes:
            liquidation_index = liquidation_indexes[-1]
            for index, cell in enumerate(row):
                if index != liquidation_index and header_role(cell) == "date":
                    row[index] = "OPER_BANCO"
            row[liquidation_index] = "FECHA LIQ"
        prepared.append(row)
    return prepared


def bbva_liquidation_dates(lines: Sequence[str]) -> list[str]:
    """Obtiene la segunda fecha de las líneas BBVA: OPER seguida de LIQ."""
    dates: list[str] = []
    for raw_line in lines:
        line = clean_text(raw_line)
        matches = list(DATE_ANY_RE.finditer(line))
        if len(matches) < 2 or matches[0].start() > 8:
            continue
        # Exige un importe después de LIQ para excluir periodos y encabezados.
        if not MONEY_RE.search(line[matches[1].end():]):
            continue
        try:
            dates.append(normalize_day_month(matches[1].group("date")))
        except (ValueError, OverflowError):
            continue
    return dates


def apply_bbva_liquidation_dates(frame: pd.DataFrame, lines: Sequence[str]) -> pd.DataFrame:
    """Reemplaza OPER por LIQ y conserva DD/MM sin inferir ningún año."""
    if frame.empty:
        return frame
    result = frame.copy()
    line_dates = bbva_liquidation_dates(lines)
    use_line_sequence = len(line_dates) == len(result)

    for index in range(len(result)):
        description = clean_text(result.at[index, "Concepto / Descripción"])
        leading_liquidation = DATE_RE.match(description)
        chosen_date: str | None = line_dates[index] if use_line_sequence else None
        if leading_liquidation:
            try:
                chosen_date = normalize_day_month(leading_liquidation.group("date"))
                result.at[index, "Concepto / Descripción"] = clean_text(
                    description[leading_liquidation.end():]
                )
            except (ValueError, OverflowError):
                pass
        if chosen_date is None:
            # En tablas con columnas separadas, prepare_bbva_liquidation_rows ya
            # hizo que la fecha existente provenga de LIQ; aquí sólo quitamos el año.
            try:
                chosen_date = normalize_day_month(str(result.at[index, "Fecha"]))
            except (ValueError, OverflowError):
                continue
        result.at[index, "Fecha"] = chosen_date
    return result


def parse_bbva(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    prepared_rows = prepare_bbva_liquidation_rows(table_rows)
    frame = parse_bank(lines, prepared_rows, CONFIGS["BBVA"])
    return apply_bbva_liquidation_dates(frame, lines)


def parse_banorte(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banorte"])


def parse_santander(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    frame = parse_bank(lines, table_rows, CONFIGS["Santander"])
    return remove_duplicated_declared_totals(frame, lines)


def parse_citibanamex(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Citibanamex"])


def parse_scotiabank(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Scotiabank"])


def parse_hsbc(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["HSBC"])


def parse_afirme(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banca Afirme"])


def parse_banregio(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banregio"])


def parse_inbursa(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco Inbursa"])


def parse_banco_azteca(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco Azteca"])


def parse_banbajio(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco del Bajío"])


def parse_bancoppel(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["BanCoppel"])


def parse_bancrea(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco Bancrea"])


def parse_mifel(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco Mifel"])


def parse_actinver(lines: Sequence[str], table_rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    return parse_bank(lines, table_rows, CONFIGS["Banco Actinver"])


PARSERS: dict[str, Callable[[Sequence[str], Sequence[Sequence[str]]], pd.DataFrame]] = {
    "BBVA": parse_bbva,
    "Banorte": parse_banorte,
    "Santander": parse_santander,
    "Citibanamex": parse_citibanamex,
    "Scotiabank": parse_scotiabank,
    "HSBC": parse_hsbc,
    "Banca Afirme": parse_afirme,
    "Banregio": parse_banregio,
    "Banco Inbursa": parse_inbursa,
    "Banco Azteca": parse_banco_azteca,
    "Banco del Bajío": parse_banbajio,
    "BanCoppel": parse_bancoppel,
    "Banco Bancrea": parse_bancrea,
    "Banco Mifel": parse_mifel,
    "Banco Actinver": parse_actinver,
}


def dataframe_to_excel(frame: pd.DataFrame) -> bytes:
    """Genera un XLSX con filtros, encabezados, formatos y anchos adecuados."""
    output = io.BytesIO()
    export = frame.copy()

    def excel_date(value: object) -> object:
        text = clean_text(value)
        if re.fullmatch(r"\d{2}/\d{2}", text):
            return text
        try:
            return datetime.strptime(text, "%d/%m/%Y")
        except ValueError:
            return text

    export["Fecha"] = export["Fecha"].map(excel_date)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name="Movimientos", index=False)
        sheet = writer.book["Movimientos"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for cell in sheet["A"][1:]:
            cell.number_format = "DD/MM/YYYY" if isinstance(cell.value, datetime) else "@"
        for col in ("C", "D", "E"):
            for cell in sheet[col][1:]:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
        for index, column_cells in enumerate(sheet.columns, start=1):
            values = [len(str(cell.value or "")) for cell in column_cells]
            width = min(max(max(values, default=10) + 2, 12), 60)
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.column_dimensions["B"].width = min(max(sheet.column_dimensions["B"].width, 35), 80)
    return output.getvalue()


def secret_value(*keys: str, default: object = None) -> object:
    """Lee claves anidadas de st.secrets sin asumir que la seccion existe."""
    try:
        value: object = st.secrets
        for key in keys:
            value = value[key]  # type: ignore[index]
        return value
    except Exception:
        return default


def smtp_setting(name: str, default: object = None) -> object:
    """Obtiene SMTP desde Streamlit Secrets o variables seguras del servidor."""
    environment_value = os.getenv(f"SMTP_{name.upper()}")
    if environment_value is not None:
        return environment_value
    secret = secret_value("smtp", name.lower(), default=None)
    if secret is not None:
        return secret
    return default


def as_bool(value: object) -> bool:
    """Convierte valores booleanos de TOML o variables de entorno."""
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "si", "sí", "on"}


def smtp_configuration() -> tuple[str, int, str, str, str, bool]:
    """Devuelve la configuración SMTP validada del servidor."""
    host = str(smtp_setting("host", "smtp.gmail.com"))
    port = int(smtp_setting("port", 587))
    username = str(smtp_setting("username", ""))
    password = str(smtp_setting("password", ""))
    sender = str(smtp_setting("sender", username))
    use_ssl = as_bool(smtp_setting("use_ssl", False))
    if not username or not password or not sender:
        raise RuntimeError("Faltan las credenciales seguras de SMTP")
    return host, port, username, password, sender, use_ssl


def send_smtp_message(message: EmailMessage) -> None:
    """Envía un mensaje usando la configuración SMTP segura."""
    host, port, username, password, _sender, use_ssl = smtp_configuration()
    context = ssl.create_default_context()
    if use_ssl:
        with smtplib.SMTP_SSL(host, port, context=context, timeout=30) as server:
            server.login(username, password)
            server.send_message(message)
    else:
        with smtplib.SMTP(host, port, timeout=30) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(username, password)
            server.send_message(message)


def send_excel_email(recipient: str, excel_bytes: bytes, bank: str) -> None:
    """Envía el resultado al usuario sólo cuando éste solicita el correo."""
    _host, _port, _username, _password, sender, _use_ssl = smtp_configuration()

    message = EmailMessage()
    message["Subject"] = f"Estado de cuenta {bank} convertido a Excel"
    message["From"] = sender
    message["To"] = recipient
    message.set_content("Adjuntamos los movimientos extraidos de tu estado de cuenta.")
    message.add_attachment(
        excel_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"movimientos_{bank.lower()}.xlsx",
    )
    send_smtp_message(message)


def send_usage_notification(user_email: str, bank: str, movement_count: int) -> None:
    """Registra por email una conversión sin adjuntar información bancaria."""
    _host, _port, _username, _password, sender, _use_ssl = smtp_configuration()
    timestamp = datetime.now().astimezone().strftime("%d/%m/%Y %H:%M:%S %Z")
    message = EmailMessage()
    message["Subject"] = f"Nueva conversión en EBC Fiscal - {bank}"
    message["From"] = sender
    message["To"] = sender
    message.set_content(
        "Se realizó una nueva conversión en la herramienta EBC Fiscal.\n\n"
        f"Correo del usuario: {user_email}\n"
        f"Banco seleccionado: {bank}\n"
        f"Movimientos detectados: {movement_count:,}\n"
        f"Fecha y hora: {timestamp}\n\n"
        "Por seguridad, este aviso no contiene ni adjunta información del estado de cuenta."
    )
    send_smtp_message(message)


def merge_preview_edits(original: pd.DataFrame, edited_preview: pd.DataFrame) -> pd.DataFrame:
    result = original.copy()
    edited = edited_preview.reindex(columns=COLUMNS).copy()
    for column in ("Depósito", "Retiro", "Saldo"):
        edited[column] = pd.to_numeric(edited[column], errors="coerce")
    result.iloc[: len(edited)] = edited.to_numpy()
    return result


def inject_styles() -> None:
    """Aplica una identidad visual limpia sin dependencias de frontend."""
    st.markdown(
        """
        <style>
        :root {
            --ink: #102a43;
            --muted: #627d98;
            --brand: #0f766e;
            --brand-dark: #115e59;
            --surface: rgba(255, 255, 255, 0.94);
            --line: #d9e2ec;
        }

        .stApp {
            background:
                radial-gradient(circle at 8% 0%, rgba(45, 212, 191, .14), transparent 28rem),
                radial-gradient(circle at 96% 12%, rgba(59, 130, 246, .10), transparent 26rem),
                #f5f8fb;
            color: var(--ink);
        }

        [data-testid="stHeader"] { background: transparent; }
        [data-testid="stToolbar"] { right: 1rem; }
        [data-testid="stAppViewContainer"] > .main .block-container {
            max-width: 1120px;
            padding-top: 2rem;
            padding-bottom: 4rem;
        }

        .hero {
            position: relative;
            overflow: hidden;
            padding: 2.7rem 3rem;
            margin-bottom: 1.4rem;
            border: 1px solid rgba(255,255,255,.14);
            border-radius: 28px;
            background: linear-gradient(125deg, #102a43 0%, #123f56 55%, #0f766e 120%);
            box-shadow: 0 24px 60px rgba(16, 42, 67, .18);
            color: white;
        }

        .hero::after {
            content: "";
            position: absolute;
            width: 240px;
            height: 240px;
            right: -55px;
            top: -75px;
            border-radius: 50%;
            background: rgba(45, 212, 191, .16);
            box-shadow: 0 0 0 48px rgba(45, 212, 191, .055);
        }

        .hero-badge {
            display: inline-flex;
            align-items: center;
            gap: .45rem;
            padding: .38rem .72rem;
            border: 1px solid rgba(255,255,255,.22);
            border-radius: 999px;
            background: rgba(255,255,255,.1);
            color: #ccfbf1;
            font-size: .76rem;
            font-weight: 700;
            letter-spacing: .08em;
            text-transform: uppercase;
        }

        .hero h1 {
            max-width: 760px;
            margin: 1rem 0 .55rem;
            color: white;
            font-size: clamp(2rem, 4vw, 3.35rem);
            line-height: 1.04;
            letter-spacing: -.045em;
        }

        .hero p {
            max-width: 680px;
            margin: 0;
            color: #d9eaf2;
            font-size: 1.05rem;
            line-height: 1.65;
        }

        .section-title {
            margin: .2rem 0 -.35rem;
            color: var(--ink);
            font-size: 1.25rem;
            font-weight: 750;
            letter-spacing: -.02em;
        }

        .section-copy {
            margin-bottom: .75rem;
            color: var(--muted);
            font-size: .92rem;
        }

        [data-testid="stForm"] {
            padding: 1.7rem 1.8rem 1.35rem;
            border: 1px solid rgba(188, 204, 220, .75);
            border-radius: 22px;
            background: var(--surface);
            box-shadow: 0 14px 38px rgba(50, 73, 94, .08);
        }

        [data-testid="stFileUploaderDropzone"] {
            min-height: 120px;
            border: 1.5px dashed #9fb3c8;
            border-radius: 15px;
            background: #f8fbfd;
        }

        [data-testid="stTextInput"] input,
        [data-baseweb="select"] > div {
            border-color: #bcccdc;
            border-radius: 11px;
            background: #fbfdff;
        }

        .stButton > button,
        .stDownloadButton > button,
        [data-testid="stFormSubmitButton"] > button {
            min-height: 3rem;
            border: 0;
            border-radius: 12px;
            background: linear-gradient(135deg, var(--brand), #0d9488);
            box-shadow: 0 8px 20px rgba(15, 118, 110, .2);
            color: white;
            font-weight: 750;
            transition: transform .16s ease, box-shadow .16s ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        [data-testid="stFormSubmitButton"] > button:hover {
            transform: translateY(-1px);
            box-shadow: 0 11px 24px rgba(15, 118, 110, .28);
            color: white;
        }

        [data-testid="stMetric"] {
            padding: 1rem 1.15rem;
            border: 1px solid #d9e2ec;
            border-radius: 16px;
            background: rgba(255,255,255,.88);
            box-shadow: 0 8px 24px rgba(50, 73, 94, .055);
        }

        [data-testid="stDataFrame"] {
            overflow: hidden;
            border: 1px solid #d9e2ec;
            border-radius: 16px;
            box-shadow: 0 8px 24px rgba(50, 73, 94, .055);
        }

        [data-testid="stAlert"] { border-radius: 13px; }
        details { border-radius: 14px !important; }

        .privacy-note {
            margin: .85rem 0 0;
            color: #829ab1;
            font-size: .78rem;
            text-align: center;
        }

        .app-footer {
            padding-top: 2rem;
            color: #829ab1;
            font-size: .78rem;
            text-align: center;
        }

        @media (max-width: 720px) {
            [data-testid="stAppViewContainer"] > .main .block-container { padding: .8rem .8rem 2rem; }
            .hero { padding: 2rem 1.35rem; border-radius: 20px; }
            .hero h1 { font-size: 2.15rem; }
            [data-testid="stForm"] { padding: 1.2rem 1rem 1rem; border-radius: 17px; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_app() -> None:
    st.set_page_config(
        page_title="Estado de cuenta a Excel",
        page_icon="↗",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    inject_styles()
    st.markdown(
        """
        <section class="hero">
            <span class="hero-badge">✦ Conversión inteligente</span>
            <h1>Tu estado de cuenta,<br>listo para trabajar.</h1>
            <p>Convierte movimientos bancarios de PDF a Excel en segundos. Revisa, corrige y descarga un archivo limpio y ordenado.</p>
        </section>
        """,
        unsafe_allow_html=True,
    )

    with st.form("conversion_form"):
        st.markdown('<div class="section-title">Comienza la conversión</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="section-copy">Completa los datos y selecciona tu estado de cuenta en PDF.</div>',
            unsafe_allow_html=True,
        )
        contact_col, bank_col = st.columns(2, gap="large")
        with contact_col:
            email = st.text_input(
                "Correo electrónico",
                placeholder="nombre@ejemplo.com",
                help="La dirección se registra al realizar una conversión.",
            )
        with bank_col:
            bank = st.selectbox("Banco", BANKS)
        pdf_file = st.file_uploader("Archivo PDF", type=["pdf"], accept_multiple_files=False)
        send_automatically = st.checkbox("Enviar también el Excel por email")
        submitted = st.form_submit_button("Convertir a Excel  →", type="primary", use_container_width=True)
        st.markdown(
            '<div class="privacy-note">🔒 Tu correo se registra para identificar la solicitud. '
            'El Excel sólo se envía si marcas la casilla y el documento se procesa únicamente para convertirlo.</div>',
            unsafe_allow_html=True,
        )

    if submitted:
        st.session_state.pop("conversion", None)
        if not EMAIL_RE.fullmatch(email.strip()):
            st.error("Ingresa un email válido.")
        elif pdf_file is None:
            st.error("Selecciona un archivo PDF.")
        elif pdf_file.type not in {"application/pdf", "application/x-pdf"} and not pdf_file.name.lower().endswith(".pdf"):
            st.error("El archivo debe ser un PDF.")
        else:
            try:
                pdf_bytes = pdf_file.getvalue()
                if not pdf_bytes.startswith(b"%PDF"):
                    raise ValueError("El archivo no tiene una firma PDF válida.")
                with st.spinner("Extrayendo y organizando movimientos…"):
                    lines, tables = extract_pdf_rows(pdf_bytes)
                    frame = PARSERS[bank](lines, tables)
                if frame.empty:
                    if lines:
                        st.warning(
                            "El PDF sí contiene texto, pero no se encontraron movimientos monetarios "
                            "que puedan validarse con seguridad. Puede usar una estructura bancaria "
                            "distinta; no se generó un Excel para evitar datos incorrectos."
                        )
                        st.caption(
                            f"Diagnóstico: {len(lines):,} líneas de texto y "
                            f"{len(tables):,} filas tabulares analizadas."
                        )
                    else:
                        st.warning(
                            "No se encontró texto seleccionable. El PDF puede estar escaneado, "
                            "protegido o dañado."
                        )
                    st.session_state.pop("conversion", None)
                else:
                    st.session_state["conversion"] = {
                        "frame": frame,
                        "bank": bank,
                        "email": email.strip(),
                        "auto_send": send_automatically,
                        "sent": False,
                        "usage_notification_attempted": False,
                        "usage_notification_sent": False,
                        "validation_warnings": validate_extraction_totals(frame, lines),
                    }
                    st.success(f"Se detectaron {len(frame):,} movimientos.")
            except Exception as exc:
                st.session_state.pop("conversion", None)
                st.error(f"No fue posible procesar el PDF: {exc}")

    conversion = st.session_state.get("conversion")
    if not conversion:
        st.markdown(
            '<div class="app-footer">Compatible con 15 bancos de México · Conversión segura a Excel</div>',
            unsafe_allow_html=True,
        )
        return

    if not conversion.get("usage_notification_attempted", False):
        # Se marca antes del envío para impedir duplicados si Streamlit vuelve
        # a ejecutar la página mientras el mensaje está en tránsito.
        conversion["usage_notification_attempted"] = True
        try:
            send_usage_notification(
                conversion["email"],
                conversion["bank"],
                len(conversion["frame"]),
            )
            conversion["usage_notification_sent"] = True
        except Exception as exc:
            st.warning(f"El Excel está listo, pero no fue posible registrar la solicitud: {exc}")

    st.markdown('<div class="section-title">Vista previa editable</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-copy">Corrige cualquier dato de las primeras 10 filas antes de generar el archivo.</div>',
        unsafe_allow_html=True,
    )
    for warning in conversion.get("validation_warnings", []):
        st.warning(f"Revisión necesaria: {warning}")
    metric_a, metric_b, metric_c = st.columns(3, gap="medium")
    with metric_a:
        st.metric("Movimientos", f"{len(conversion['frame']):,}")
    with metric_b:
        st.metric("Depósitos", f"${conversion['frame']['Depósito'].sum():,.2f}")
    with metric_c:
        st.metric("Retiros", f"${conversion['frame']['Retiro'].sum():,.2f}")
    preview = conversion["frame"].head(10).copy()
    edited_preview = st.data_editor(
        preview,
        key="preview_editor",
        hide_index=True,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "Fecha": st.column_config.TextColumn("Fecha", help="DD/MM para BBVA; DD/MM/YYYY para otros bancos"),
            "Concepto / Descripción": st.column_config.TextColumn("Concepto / Descripción", width="large"),
            "Depósito": st.column_config.NumberColumn("Depósito", format="$ %.2f", min_value=0.0),
            "Retiro": st.column_config.NumberColumn("Retiro", format="$ %.2f", min_value=0.0),
            "Saldo": st.column_config.NumberColumn("Saldo", format="$ %.2f"),
        },
    )
    final_frame = merge_preview_edits(conversion["frame"], edited_preview)
    try:
        excel_bytes = dataframe_to_excel(final_frame)
    except Exception as exc:
        st.error(f"Revisa las fechas editadas; no se pudo crear el Excel: {exc}")
        return

    st.download_button(
        "Descargar Excel",
        data=excel_bytes,
        file_name=f"movimientos_{conversion['bank'].lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    if conversion["auto_send"] and not conversion["sent"]:
        try:
            send_excel_email(conversion["email"], excel_bytes, conversion["bank"])
            conversion["sent"] = True
            st.success(f"Excel enviado a {conversion['email']}.")
        except Exception as exc:
            st.warning(f"El Excel está listo, pero no se pudo enviar por email: {exc}")

    st.markdown(
        '<div class="app-footer">Conversión terminada · Revisa siempre los datos antes de utilizarlos.</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    render_app()
